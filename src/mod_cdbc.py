import requests

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from dotenv import load_dotenv
import os

cert_path = os.path.join(os.path.dirname(__file__), "cert.pem")
load_dotenv()

class cdbc:
    def get_token():
        url = f"{os.getenv('ENDPOINT_CDBC_VIETTEL')}/workflow_system/api-operator/public/token/"
        body = {"username": os.getenv('USER_CDBC'), "password": os.getenv('PASSWORD_CDBC')}
        res = requests.post(url, json=body, verify=cert_path)
        if res.status_code == 200:
            return res.json()['data']["token"]
        else:
            return "Error"
        
    def get_report(token):
        url = f"{os.getenv('ENDPOINT_CDBC_VIETTEL')}/workflow_system/api-operator/public/list_report/"
        header = {
            "Authorization": f"Bearer {token}"
        }
        res = requests.get(url, headers=header, verify=cert_path)
        if res.status_code == 200:
            data =  res.json()
            list_report = data['data']['data']
            for i in list_report:
                with open("api_response.txt", "a", encoding="utf-8") as f:
                    f.write(f"{i['report_id']}\n{i['report_name']}\n")
        else:
            return "Error"
    
    def get_layout(token, report_id):
        url = f"os.getenv('ENDPOINT_CDBC_VIETTEL')/workflow_system/api-operator/public/schema_report/?report_id={report_id}"
        header = {
            "Authorization": f"Bearer {token}"
        }
        res = requests.get(url, headers=header, verify=cert_path)
        if res.status_code == 200:
            data =  res.json()
            return data
        else:
            return "Error"
        
    def make_excel(data, file_name):

        try:
            columns = data["data"]["data"]["columns"]
            rows_data = data["data"]["data"]["rows"]
            rows = sorted(rows_data, key=lambda x: x["row_id"])

            # ===== Xử lý CỘT (columns) =====
            grouped = {}
            for col in columns:
                parent_code = col.get("parent_column_code")
                parent_name = col.get("parent_column_name")
                col_code = col.get("column_code")
                col_name = col.get("column_name")

                parent_columns = f"{parent_code} - {parent_name}" if parent_code and parent_name else None
                column = f"{col_code} - {col_name}"

                if parent_code and parent_name:
                    grouped.setdefault(parent_columns, []).append(column)
                else:
                    grouped[column] = []  # không có parent

            # ===== Xây dựng cây row (cha–con) =====
            row_dict = {r["row_id"]: r for r in rows}

            children_map = {}
            for r in rows:
                parent_id = r.get("parent_row_id")
                children_map.setdefault(parent_id, []).append(r)

            # ===== Duyệt đệ quy để tạo danh sách hàng có thụt lề =====
            def build_row_list(parent_id=None, level=0):
                items = []
                for r in children_map.get(parent_id, []):
                    prefix = " " * (level * 4)  # 4 dấu cách mỗi cấp
                    row_text = f"{prefix}{r['row_id']} - {r['row_name']}"
                    items.append(row_text)
                    # đệ quy cho con
                    items.extend(build_row_list(r["row_id"], level + 1))
                return items

            row_list = build_row_list()  # bắt đầu từ cấp gốc (parent_id=None)

            # ===== Tạo tiêu đề cột =====
            header_row_1 = []
            header_row_2 = []

            for parent, children in grouped.items():
                if children:
                    header_row_1.extend([parent] * len(children))
                    header_row_2.extend(children)
                else:
                    header_row_1.append(parent)
                    header_row_2.append("")

            # ===== Tạo DataFrame =====
            total_rows = 2 + len(row_list)
            num_columns = len(header_row_1)

            df = pd.DataFrame(index=range(total_rows), columns=range(num_columns))
            df.iloc[0] = header_row_1
            df.iloc[1] = header_row_2

            # Cột “Hàng”: 2 dòng đầu None + các hàng thật
            hang_values = [None, None] + row_list
            df.insert(0, "Hàng", hang_values)

            # ===== Xuất ra Excel =====
            excel_path = f"{file_name}.xlsx"
            df.to_excel(excel_path, index=False, header=False)

            # ===== Merge cột cha trong Excel =====
            wb = load_workbook(excel_path)
            ws = wb.active

            col_idx = 2  # vì cột 1 là “Hàng”
            while col_idx <= num_columns + 1:
                parent = header_row_1[col_idx - 2]
                start = col_idx
                while col_idx <= num_columns + 1 and header_row_1[col_idx - 2] == parent:
                    col_idx += 1
                end = col_idx - 1
                if end > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                for c in range(start, end + 1):
                    ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")

            # Căn trái cho cột “Hàng”
            for row_idx in range(3, total_rows + 1):
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            wb.save(excel_path)
            return f"💕 File được lưu thành công tại thư mục đặt tools 💕"
        except:
            return f"💀 Lỗi rồi. Liên hệ NAMNT62 để kiểm tra lỗi 💀"
